"""tests/test_revolut_parser.py — unit tests for Revolut XLSX parser"""
from __future__ import annotations

import io
import tempfile
from pathlib import Path

import openpyxl
import pytest

from regime_trader.services.revolut_parser import (
    compute_cash_balance_usd,
    net_positions_from_rows,
    parse_xlsx,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_xlsx(rows: list[tuple]) -> Path:
    """Write an in-memory XLSX with Revolut header + given rows, return tmp path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("Date", "Ticker", "Type", "Quantity", "Price per share", "Total Amount", "Currency", "FX Rate"))
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return Path(tmp.name)


def _make_cash_xlsx(rows: list[tuple]) -> io.BytesIO:
    """Write a BytesIO XLSX with the full Revolut header (including Total Amount) for cash tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("Date", "Ticker", "Type", "Quantity", "Price per share", "Total Amount", "Currency", "FX Rate"))
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Tests ─────────────────────────────────────────────────────────────────────

class TestNetPositionsFromRows:
    def test_single_buy_creates_position(self):
        rows = [
            {"ticker": "MSFT", "type": "BUY - MARKET", "qty": 2.0, "price": 393.22, "currency": "USD"},
        ]
        result = net_positions_from_rows(rows)
        assert len(result) == 1
        assert result[0]["ticker"] == "MSFT"
        assert result[0]["net_qty"] == pytest.approx(2.0)
        assert result[0]["avg_cost"] == pytest.approx(393.22)

    def test_full_sell_removes_position(self):
        rows = [
            {"ticker": "COIN", "type": "BUY - MARKET",  "qty": 5.0,  "price": 200.0, "currency": "USD"},
            {"ticker": "COIN", "type": "SELL - MARKET", "qty": 5.0,  "price": 180.0, "currency": "USD"},
        ]
        result = net_positions_from_rows(rows)
        assert result == []

    def test_partial_sell_keeps_remainder(self):
        rows = [
            {"ticker": "DDOG", "type": "BUY - MARKET",  "qty": 4.0, "price": 100.0, "currency": "USD"},
            {"ticker": "DDOG", "type": "SELL - MARKET", "qty": 1.5, "price": 110.0, "currency": "USD"},
        ]
        result = net_positions_from_rows(rows)
        assert len(result) == 1
        assert result[0]["net_qty"] == pytest.approx(2.5)

    def test_weighted_avg_cost_basis(self):
        rows = [
            {"ticker": "OXY", "type": "BUY - MARKET", "qty": 5.0, "price": 50.0, "currency": "USD"},
            {"ticker": "OXY", "type": "BUY - MARKET", "qty": 5.0, "price": 60.0, "currency": "USD"},
        ]
        result = net_positions_from_rows(rows)
        assert result[0]["avg_cost"] == pytest.approx(55.0)  # (5*50 + 5*60) / 10

    def test_dividend_rows_ignored(self):
        rows = [
            {"ticker": "ORCL", "type": "BUY - MARKET", "qty": 2.0, "price": 176.0, "currency": "USD"},
            {"ticker": "ORCL", "type": "DIVIDEND",      "qty": None, "price": None, "currency": "USD"},
        ]
        result = net_positions_from_rows(rows)
        assert len(result) == 1
        assert result[0]["net_qty"] == pytest.approx(2.0)

    def test_source_field_is_revolut(self):
        rows = [
            {"ticker": "PANW", "type": "BUY - MARKET", "qty": 3.0, "price": 169.0, "currency": "USD"},
        ]
        result = net_positions_from_rows(rows)
        assert result[0]["source"] == "revolut"


class TestParseXlsx:
    def test_parses_real_format(self):
        xlsx_path = _make_xlsx([
            ("2026-04-14T13:40:22.074Z", "NUVL", "BUY - MARKET", 2.38, "USD 104.84", "USD 250", "USD", 1.18),
            ("2026-04-14T13:40:52.225Z", "MSTR", "BUY - MARKET", 1.42, "USD 139.71", "USD 199", "USD", 1.18),
            ("2026-04-10T16:40:37.280Z", "GZF",  "SELL - MARKET", 15,  "EUR 29.23",  "EUR 438", "EUR", 1.0),
        ])
        result = parse_xlsx(xlsx_path, ticker_map={})
        tickers = {p["ticker"] for p in result}
        assert "NUVL" in tickers
        assert "MSTR" in tickers
        assert "GZF" not in tickers  # net_qty = -15 (sell with no buy)

    def test_price_string_with_currency_prefix_parsed(self):
        xlsx_path = _make_xlsx([
            ("2026-04-14T13:43:01.610Z", "MSFT", "BUY - MARKET", 1.27, "USD 393.22", "USD 500", "USD", 1.18),
        ])
        result = parse_xlsx(xlsx_path, ticker_map={})
        assert result[0]["avg_cost"] == pytest.approx(393.22, abs=0.01)

    def test_ticker_map_applied(self):
        xlsx_path = _make_xlsx([
            ("2026-04-14T13:00:00Z", "AIR1", "BUY - MARKET", 10.0, "EUR 28.0", "EUR 280", "EUR", 1.0),
        ])
        result = parse_xlsx(xlsx_path, ticker_map={"AIR1": "EADSY"})
        assert result[0]["ticker"] == "EADSY"
        assert result[0]["revolut_ticker"] == "AIR1"

    def test_cash_events_ignored(self):
        xlsx_path = _make_xlsx([
            ("2026-04-14T13:43:02Z", None, "CASH TOP-UP",    None, None, "USD 501", "USD", 1.18),
            ("2026-04-14T13:42:31Z", None, "CASH WITHDRAWAL", None, None, "EUR -458", "EUR", 1.0),
            ("2026-04-14T13:40:22Z", "OXY", "BUY - MARKET",  5.0, "USD 54.50", "USD 272", "USD", 1.18),
        ])
        result = parse_xlsx(xlsx_path, ticker_map={})
        assert len(result) == 1
        assert result[0]["ticker"] == "OXY"

    def test_raises_on_missing_header_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(("Not", "A", "Valid", "Header"))
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        with pytest.raises(ValueError, match="Could not find header row"):
            parse_xlsx(Path(tmp.name), ticker_map={})

    def test_raises_on_missing_required_column(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(("Date", "Ticker", "Type"))  # Missing Quantity and Price per share
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        with pytest.raises(ValueError, match="missing required columns"):
            parse_xlsx(Path(tmp.name), ticker_map={})


class TestComputeCashBalanceUsd:
    def test_cash_top_up_adds_to_balance(self):
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP", None, None, "USD 5000", "USD", 1.0),
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(5000.0)

    def test_buy_reduces_cash(self):
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP",  None, None, "USD 2000", "USD", 1.0),
            ("2026-01-02", "OXY", "BUY - MARKET", 10, "USD 50", "USD 500", "USD", 1.0),
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(1500.0)

    def test_sell_adds_back_to_cash(self):
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP",   None, None, "USD 1000", "USD", 1.0),
            ("2026-01-02", "OXY", "BUY - MARKET",  5, "USD 50", "USD 250", "USD", 1.0),
            ("2026-01-03", "OXY", "SELL - MARKET", 5, "USD 60", "USD 300", "USD", 1.0),
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(1050.0)

    def test_fx_conversion_applied(self):
        """EUR amount * FX rate should give USD equivalent."""
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP", None, None, "EUR 1000", "EUR", 1.10),
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(1100.0)

    def test_blank_rows_ignored(self):
        """Rows with empty Type must not affect the balance."""
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP", None, None, "USD 1000", "USD", 1.0),
            (None, None, None, None, None, "USD 999", "USD", 1.0),  # blank type row
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(1000.0)

    def test_withdrawal_already_negative(self):
        """Revolut encodes withdrawals with a negative Total Amount."""
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP",    None, None, "USD 2000", "USD", 1.0),
            ("2026-01-02", None, "CASH WITHDRAWAL", None, None, "USD -500", "USD", 1.0),
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(1500.0)

    def test_missing_total_amount_row_skipped(self):
        buf = _make_cash_xlsx([
            ("2026-01-01", None, "CASH TOP-UP", None, None, "USD 800", "USD", 1.0),
            ("2026-01-02", None, "DIVIDEND",    None, None, None,       "USD", 1.0),
        ])
        assert compute_cash_balance_usd(buf) == pytest.approx(800.0)
