"""regime_trader/services/revolut_parser.py
Parse a Revolut trading account statement (.xlsx) into net positions.

Revolut XLSX columns:
  Date | Ticker | Type | Quantity | Price per share | Total Amount | Currency | FX Rate

Transaction types handled:
  BUY - MARKET / BUY - LIMIT / BUY - STOP  → add to position
  SELL - MARKET / SELL - LIMIT / SELL - STOP → reduce position
  DIVIDEND, CASH TOP-UP, CASH WITHDRAWAL    → ignored
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

_BUY_TYPES  = {"BUY - MARKET", "BUY - LIMIT", "BUY - STOP"}
_SELL_TYPES = {"SELL - MARKET", "SELL - LIMIT", "SELL - STOP"}

# Transaction types that ADD cash (positive) or REMOVE cash (negative Total Amount already embedded)
_CASH_FLOW_TYPES = {
    "CASH TOP-UP", "CASH WITHDRAWAL",
    "SELL - MARKET", "SELL - LIMIT", "SELL - STOP",
    "DIVIDEND", "REWARD", "RETURN OF CAPITAL", "MERGER - CASH",
}

_DEFAULT_MAP = Path(__file__).parent.parent.parent / "data" / "revolut_ticker_map.json"


def _load_default_ticker_map() -> Dict[str, str]:
    if _DEFAULT_MAP.exists():
        return json.loads(_DEFAULT_MAP.read_text(encoding="utf-8"))
    return {}


def _parse_price(raw: Any) -> float:
    """Parse price field which may be 'USD 104.84', 'EUR 29.23', or a bare float."""
    if raw is None:
        return 0.0
    parts = str(raw).strip().split()
    try:
        return float(parts[-1].replace(",", ""))
    except (ValueError, IndexError):
        return 0.0


def net_positions_from_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Compute net positions from a list of normalised transaction dicts.

    Each dict must have: ticker, type, qty (float|None), price (float|None), currency.
    Returns only positions with net_qty > 1e-6.
    """
    buys:    Dict[str, List[tuple[float, float]]] = {}
    net_qty: Dict[str, float] = {}
    currency_map: Dict[str, str] = {}

    for row in rows:
        tx_type = str(row.get("type", "")).strip()
        ticker  = row.get("ticker")
        qty     = row.get("qty")
        price_raw = row.get("price")
        price = float(price_raw) if price_raw is not None else 0.0
        currency = str(row.get("currency", "USD")).strip()

        if not ticker or qty is None:
            continue

        qty = float(qty)

        if tx_type in _BUY_TYPES:
            net_qty[ticker] = net_qty.get(ticker, 0.0) + qty
            buys.setdefault(ticker, []).append((qty, float(price)))
            currency_map[ticker] = currency
        elif tx_type in _SELL_TYPES:
            net_qty[ticker] = net_qty.get(ticker, 0.0) - qty

    positions = []
    for ticker, remaining in net_qty.items():
        if remaining <= 1e-6:
            continue
        buy_list = buys.get(ticker, [])
        total_qty_bought = sum(q for q, _ in buy_list)
        total_cost = sum(q * p for q, p in buy_list)
        avg_cost = total_cost / total_qty_bought if total_qty_bought > 0 else 0.0
        positions.append({
            "ticker":          ticker,
            "revolut_ticker":  ticker,
            "net_qty":         round(remaining, 8),
            "avg_cost":        round(avg_cost, 4),
            "currency":        currency_map.get(ticker, "USD"),
            "source":          "revolut",
        })

    return sorted(positions, key=lambda p: p["ticker"])


def compute_cash_balance_usd(filepath: str | Path) -> float:
    """Compute net uninvested cash balance in USD from a full Revolut trading XLSX.

    Logic:
      CASH TOP-UP / SELL / DIVIDEND / REWARD etc. → add amount (income or inflow)
      BUY → subtract amount (Total Amount is positive but represents outflow)
      CASH WITHDRAWAL → add amount (already negative in Total Amount field)
      Non-USD amounts are converted to USD using the stored FX Rate column.

    Returns the net cash balance in USD (may be approximate due to FX rounding).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers: Optional[List[str]] = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0]).strip() == "Date":
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if headers is None:
        return 0.0

    col = {h: i for i, h in enumerate(headers)}
    if "Total Amount" not in col:
        return 0.0

    cash_usd = 0.0
    for raw in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        tx_type   = str(raw[col["Type"]]).strip() if "Type" in col and raw[col["Type"]] is not None else ""
        total_raw = raw[col["Total Amount"]]
        currency  = str(raw[col["Currency"]]).strip() if "Currency" in col and raw[col["Currency"]] else "USD"
        fx_raw    = raw[col["FX Rate"]] if "FX Rate" in col else None
        fx_rate   = float(fx_raw) if fx_raw is not None else 1.0

        if not tx_type or total_raw is None:
            continue
        parts = str(total_raw).strip().split()
        try:
            amount = float(parts[-1].replace(",", ""))
        except (ValueError, IndexError):
            continue

        # Convert non-USD to USD using the stored FX Rate
        amount_usd = amount * fx_rate if currency != "USD" else amount

        if tx_type in _BUY_TYPES:
            cash_usd -= amount_usd  # BUY: data shows positive, but it is cash OUT
        elif tx_type in _CASH_FLOW_TYPES:
            cash_usd += amount_usd  # WITHDRAWAL already carries negative sign

    return round(cash_usd, 2)


def parse_xlsx(
    filepath: str | Path,
    ticker_map: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """Parse a Revolut XLSX statement into a list of net positions.

    Args:
        filepath:   Path to the .xlsx file.
        ticker_map: Optional {revolut_symbol: universe_symbol} dict.
                    Defaults to data/revolut_ticker_map.json.

    Returns:
        List of position dicts sorted by ticker.
    """
    if ticker_map is None:
        ticker_map = _load_default_ticker_map()

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Locate the header row (first row where col 0 == "Date")
    headers: Optional[List[str]] = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0]).strip() == "Date":
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if headers is None:
        raise ValueError(f"Could not find header row in {filepath}")

    col = {h: i for i, h in enumerate(headers)}

    _REQUIRED_COLS = {"Ticker", "Type", "Quantity", "Price per share"}
    missing = _REQUIRED_COLS - col.keys()
    if missing:
        raise ValueError(f"XLSX missing required columns: {missing}")

    rows = []
    for raw in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        rows.append({
            "ticker":   str(raw[col["Ticker"]]).strip() if raw[col["Ticker"]] else None,
            "type":     str(raw[col["Type"]]).strip()   if raw[col["Type"]]   else "",
            "qty":      raw[col["Quantity"]],
            "price":    _parse_price(raw[col["Price per share"]]),
            "currency": str(raw[col["Currency"]]).strip() if "Currency" in col and raw[col["Currency"]] else "USD",
        })

    positions = net_positions_from_rows(rows)

    # Apply ticker mapping
    for pos in positions:
        original = pos["ticker"]
        mapped   = ticker_map.get(original, original)
        pos["ticker"]          = mapped
        pos["revolut_ticker"]  = original

    return positions
