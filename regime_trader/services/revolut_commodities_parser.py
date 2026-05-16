"""regime_trader/services/revolut_commodities_parser.py
Parse a Revolut commodities/metals account statement (.xlsx) into net holdings.

Revolut metals statement columns:
  Type | Product | Started Date | Completed Date | Description | Amount | Fee | Currency | State | Balance

Logic:
  - Group rows by Currency (XAU = gold, XAG = silver, etc.)
  - Filter to State == "COMPLETED" rows only
  - Net holding = last Balance value for each commodity currency
  - Average cost = total positive amount converted via GBP/USD exchange (approx)

Commodity → yfinance ticker map:
  XAU → GC=F  (COMEX Gold futures, price in USD/oz)
  XAG → SI=F  (COMEX Silver futures, price in USD/oz)
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

log = logging.getLogger(__name__)

_COMMODITY_META: Dict[str, Dict[str, str]] = {
    "XAU": {"name": "Gold",   "unit": "oz", "yf_ticker": "GC=F"},
    "XAG": {"name": "Silver", "unit": "oz", "yf_ticker": "SI=F"},
    "XPT": {"name": "Platinum", "unit": "oz", "yf_ticker": "PL=F"},
    "XPD": {"name": "Palladium", "unit": "oz", "yf_ticker": "PA=F"},
}

# Commodity currency codes that indicate a metals holding (not fiat)
_COMMODITY_CURRENCIES = set(_COMMODITY_META.keys())


def parse_commodities_xlsx(filepath: str | Path) -> List[Dict[str, Any]]:
    """Parse a Revolut metals account statement XLSX into net commodity positions.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        List of position dicts, one per commodity currency with a non-zero balance:
        {
            "commodity": "XAU",
            "name": "Gold",
            "amount": 0.4632,       # oz (or other unit)
            "unit": "oz",
            "yf_ticker": "GC=F",    # yfinance symbol for live price
            "source": "revolut_commodities",
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Locate the header row (first row where col 0 is "Type")
    headers: Optional[List[str]] = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0]).strip() == "Type":
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if headers is None:
        raise ValueError(f"Could not find header row in {filepath}")

    col = {h: i for i, h in enumerate(headers)}

    _REQUIRED = {"Type", "Currency", "State", "Balance"}
    missing = _REQUIRED - col.keys()
    if missing:
        raise ValueError(f"XLSX missing required columns: {missing}")

    # Build per-currency balance from the most-recent COMPLETED row.
    # "Completed Date" column is used for tie-breaking when present; otherwise
    # row order determines recency (last row wins). This is robust to both
    # ascending and descending chronological exports.
    date_col = col.get("Completed Date")
    balances: Dict[str, float] = {}
    balance_dates: Dict[str, str] = {}   # currency -> latest date string seen

    for raw in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        currency = str(raw[col["Currency"]]).strip() if raw[col["Currency"]] else ""
        state    = str(raw[col["State"]]).strip()    if raw[col["State"]]    else ""
        balance  = raw[col["Balance"]]

        if currency not in _COMMODITY_CURRENCIES:
            continue
        if state != "COMPLETED":
            continue
        if balance is None:
            continue

        try:
            bal_float = float(balance)
        except (TypeError, ValueError):
            continue

        if date_col is not None:
            row_date = str(raw[date_col]).strip() if raw[date_col] is not None else ""
            prev_date = balance_dates.get(currency, "")
            if row_date >= prev_date:   # lexicographic ISO-8601 sort is correct
                balances[currency] = bal_float
                balance_dates[currency] = row_date
        else:
            balances[currency] = bal_float   # last row wins when no date column

    positions = []
    for currency, amount in balances.items():
        if amount <= 1e-8:
            continue
        meta = _COMMODITY_META.get(currency, {
            "name": currency, "unit": "oz", "yf_ticker": None,
        })
        positions.append({
            "commodity":  currency,
            "name":       meta["name"],
            "amount":     round(amount, 8),
            "unit":       meta["unit"],
            "yf_ticker":  meta.get("yf_ticker"),
            "source":     "revolut_commodities",
        })

    return sorted(positions, key=lambda p: p["commodity"])
